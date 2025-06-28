import openpyxl
from openpyxl import load_workbook
import os
from datetime import datetime
import re

# Hàm tính tuổi
def calculate_age(birth_date):
    today = datetime.today()
    age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
    return age

# Hàm kiểm tra email hợp lệ
def validate_email(email):
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return bool(re.match(pattern, email))

# Hàm kiểm tra số điện thoại hợp lệ (10 hoặc 11 số)
def validate_phone(phone):
    pattern = r'^\d{10,11}$'
    return bool(re.match(pattern, phone))

# Hàm xác định mã trạng thái
def get_status_code(age, job, is_student):
    if is_student:
        return 1
    if age >= 18:
        return 3 if job else 2
    return None

# Hàm chính
def main():
    file_name = "user_info.xlsx"

    # Nếu file đã tồn tại thì mở ra, không thì tạo mới
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "User Information"

        # Header
        headers = ["Họ và tên", "Ngày sinh", "Email", "Số điện thoại", "Công việc",
                    "Tình trạng hôn nhân", "Trạng thái tuổi", "Mã trạng thái"]
        ws.append(headers)

    while True:
        print("\n--- Nhập thông tin người dùng ---")

        # Nhập họ tên
        name = input("Họ và tên: ").strip()

        # Kiểm tra ngày sinh hợp lệ
        while True:
            birth_input = input("Ngày sinh (YYYY-MM-DD): ").strip()
            try:
                birth_date = datetime.strptime(birth_input, "%Y-%m-%d")
                age = calculate_age(birth_date)
                break
            except:
                print("Ngày sinh không hợp lệ! Định dạng phải là YYYY-MM-DD.")

        # Kiểm tra email hợp lệ
        while True:
            email = input("Email: ").strip()
            if validate_email(email):
                break
            print("Email không hợp lệ!")

        # Kiểm tra số điện thoại hợp lệ
        while True:
            phone = input("Số điện thoại: ").strip()
            if validate_phone(phone):
                break
            print("Số điện thoại không hợp lệ! Phải là 10 hoặc 11 số.")

        # Nhập công việc
        job = input("Công việc (bỏ trống nếu không có): ").strip()
        if job == "":
            job = "Thất nghiệp"  

        # Nhập tình trạng hôn nhân
        married = input("Tình trạng hôn nhân (độc thân/đã kết hôn): ").strip()

        # Kiểm tra sinh viên
        student_input = input("Là sinh viên? (y/n): ").strip().lower()
        is_student = student_input == 'y'

        # Xác định trạng thái tuổi và mã trạng thái
        age_status = "Trên 18" if age >= 18 else "Dưới 18"
        status_code = get_status_code(age, job, is_student)

        # Ghi dữ liệu vào Excel
        ws.append([name, birth_input, email, phone, job, married, age_status, status_code])
        print("Thêm dữ liệu thành công.")

        # Hỏi tiếp tục hay không
        cont = input("Bạn có muốn nhập tiếp không? (y/n): ").strip().lower()
        if cont != 'y':
            break

    # Lưu file Excel
    wb.save("user_info.xlsx")
    print("Đã lưu file thành công")

# Gọi hàm chính
main()
